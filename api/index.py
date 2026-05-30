from flask import Flask, request, jsonify, send_file, render_template
import openpyxl
from openpyxl.drawing.image import Image
import qrcode
import tempfile
import os
import shutil
from io import BytesIO
import base64

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/generate', methods=['POST'])
def generate_qr():
    if 'file' not in request.files:
        return jsonify({'error': 'Không có file được upload'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'Chưa chọn file'}), 400
    
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Chỉ chấp nhận file Excel (.xlsx)'}), 400
    
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Lưu file upload
        input_path = os.path.join(temp_dir, 'input.xlsx')
        file.save(input_path)
        
        # Xử lý Excel
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        
        qr_column = "Q"
        ws.column_dimensions[qr_column].width = 18
        
        count = 0
        
        for row in range(2, ws.max_row + 1):
            value = ws[f"E{row}"].value
            
            if value is None:
                continue
            
            # Xử lý lỗi 51919.0 => 51919
            if isinstance(value, float):
                qr_text = str(int(value))
            else:
                qr_text = str(value).strip()
            
            if qr_text == "":
                continue
            
            # Tạo QR
            img = qrcode.make(qr_text)
            img_path = os.path.join(temp_dir, f"qr_{row}.png")
            img.save(img_path)
            
            excel_img = Image(img_path)
            excel_img.width = 90
            excel_img.height = 90
            
            ws.add_image(excel_img, f"{qr_column}{row}")
            ws.row_dimensions[row].height = 70
            
            count += 1
        
        # Lưu file kết quả
        output_path = os.path.join(temp_dir, 'output.xlsx')
        wb.save(output_path)
        
        # Đọc file để trả về
        with open(output_path, 'rb') as f:
            output_data = BytesIO(f.read())
        
        output_data.seek(0)
        
        return send_file(
            output_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='qr_generated.xlsx'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

@app.route('/api/generate-single', methods=['POST'])
def generate_single_qr():
    """API để tạo QR code đơn lẻ"""
    data = request.get_json()
    
    if not data or 'text' not in data:
        return jsonify({'error': 'Thiếu dữ liệu text'}), 400
    
    text = data['text'].strip()
    
    if not text:
        return jsonify({'error': 'Text không được để trống'}), 400
    
    try:
        # Tạo QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(text)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Convert to base64
        buffered = BytesIO()
        img.save(buffered, format="PNG")
        img_str = base64.b64encode(buffered.getvalue()).decode()
        
        return jsonify({
            'success': True,
            'image': f'data:image/png;base64,{img_str}'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
