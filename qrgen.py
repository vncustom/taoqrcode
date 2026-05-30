import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.drawing.image import Image
import qrcode
import tempfile
import os
import shutil


class QRGeneratorApp:

    def __init__(self, root):
        self.root = root
        self.root.title("Excel QR Generator")
        self.root.geometry("550x250")
        self.root.resizable(False, False)

        self.file_path = ""

        title = tk.Label(
            root,
            text="Tạo QR Code từ cột E và chèn vào cột Q",
            font=("Arial", 13, "bold")
        )
        title.pack(pady=15)

        self.lbl_file = tk.Label(
            root,
            text="Chưa chọn file Excel",
            wraplength=500
        )
        self.lbl_file.pack(pady=5)

        btn_browse = tk.Button(
            root,
            text="Chọn File Excel",
            width=25,
            command=self.browse_file
        )
        btn_browse.pack(pady=10)

        btn_generate = tk.Button(
            root,
            text="Tạo QR Code",
            width=25,
            command=self.generate_qr
        )
        btn_generate.pack(pady=10)

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[("Excel Files", "*.xlsx")]
        )

        if file_path:
            self.file_path = file_path
            self.lbl_file.config(text=file_path)

    def generate_qr(self):

        if not self.file_path:
            messagebox.showwarning(
                "Thông báo",
                "Vui lòng chọn file Excel"
            )
            return

        temp_dir = tempfile.mkdtemp()

        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active

            qr_column = "Q"

            # Tăng chiều rộng cột Q
            ws.column_dimensions[qr_column].width = 18

            count = 0

            for row in range(2, ws.max_row + 1):

                value = ws[f"E{row}"].value

                if value is None:
                    continue

                # Xử lý lỗi 51919.0 => 51919
                if isinstance(value, float):
                    qr_text = str(int(value))
                else:
                    qr_text = str(value).strip()

                if qr_text == "":
                    continue

                # Tạo QR
                img = qrcode.make(qr_text)

                img_path = os.path.join(
                    temp_dir,
                    f"qr_{row}.png"
                )

                img.save(img_path)

                excel_img = Image(img_path)

                excel_img.width = 90
                excel_img.height = 90

                ws.add_image(
                    excel_img,
                    f"{qr_column}{row}"
                )

                ws.row_dimensions[row].height = 70

                count += 1

            output_file = filedialog.asksaveasfilename(
                title="Lưu file kết quả",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")]
            )

            if not output_file:
                return

            wb.save(output_file)

            messagebox.showinfo(
                "Hoàn thành",
                f"Tạo thành công {count} QR Code\n\n{output_file}"
            )

        except Exception as e:
            messagebox.showerror(
                "Lỗi",
                str(e)
            )

        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    root = tk.Tk()
    app = QRGeneratorApp(root)
    root.mainloop()